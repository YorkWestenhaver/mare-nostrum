#!/usr/bin/env python3
"""
Process all downloaded raw datasets into JavaScript arrays for datasets.js.
Only uses REAL downloaded data — no fabrication.
"""
import json
import csv
import os
import openpyxl
from collections import Counter, defaultdict

RAW = os.path.join(os.path.dirname(__file__), 'raw')
SVELTE_DATA = os.path.join(os.path.dirname(__file__), '..', 'svelte-app', 'src', 'lib', 'data')


def write_json(filename, data):
    """Write processed data to a JSON file in the svelte data directory."""
    path = os.path.join(SVELTE_DATA, filename)
    with open(path, 'w') as f:
        json.dump(data, f, ensure_ascii=False)
    print(f"  → Wrote {path} ({len(data)} entries)")
    return data


# =============================================================================
# 1. McConnell 2018 — Lead emissions (UPGRADE)
#    Source: pnas.1721818115.sd01.xlsx, Sheet "Fig. 3"
#    Columns: Year Before 1950, non-background Lead Flux, 11-y median filtered Lead Emissions (kt/a)
# =============================================================================
def process_mcconnell():
    wb = openpyxl.load_workbook(os.path.join(RAW, 'pnas.1721818115.sd01.xlsx'), read_only=True)
    ws = wb['Fig. 3']
    rows = list(ws.iter_rows(values_only=True))

    data = []
    for row in rows[7:]:  # skip header rows
        bp, flux, emissions = row[0], row[1], row[2]
        if bp is None or emissions is None:
            continue
        if emissions == -0.999:
            continue
        year_ce = 1950 - bp
        if -1200 <= year_ce <= 800:
            data.append({'year': round(year_ce, 1), 'emissions': round(emissions, 4)})

    data.sort(key=lambda d: d['year'])

    # Downsample to ~5-year resolution for reasonable chart size
    # Use 5-year bins with mean
    binned = {}
    for d in data:
        bin_year = int(round(d['year'] / 5) * 5)
        if bin_year not in binned:
            binned[bin_year] = []
        binned[bin_year].append(d['emissions'])

    result = []
    for year in sorted(binned.keys()):
        vals = binned[year]
        mean_em = sum(vals) / len(vals)
        result.append({'year': year, 'emissions': round(mean_em, 4)})

    print(f"McConnell lead: {len(data)} annual points → {len(result)} 5-year bins")
    return write_json('lead_mcconnell.json', result)


# =============================================================================
# 2. Büntgen 2011 — Tree ring temperature anomaly (UPGRADE)
#    Source: buentgen2011europe.txt from NOAA NCEI
#    Column 6 (0-indexed col 5): JJA Temperature Anomaly
# =============================================================================
def process_buentgen():
    filepath = os.path.join(RAW, 'buentgen2011europe.txt')
    data = []
    with open(filepath) as f:
        lines = f.readlines()

    in_data = False
    for line in lines:
        stripped = line.strip()
        if stripped.startswith('Year') and 'TempJJA' in stripped:
            in_data = True
            continue
        if in_data:
            parts = stripped.split()
            if not parts:
                break
            try:
                year = int(parts[0])
            except ValueError:
                break
            if year > 1000:
                break
            # Column layout varies by row:
            #   4 cols: year, tempJJA, temp-, temp+  (early years, no precip)
            #   7 cols: year, precip, p-, p+, tempJJA, temp-, temp+
            #   8 cols: year, precip, p-, p+, precip2010, tempJJA, temp-, temp+
            #   9 cols: year, precip, p-, p+, precip2010, tempJJA, temp-, temp+, temp2006
            temp = None
            if len(parts) == 4:
                temp = float(parts[1])
            elif len(parts) == 7:
                temp = float(parts[4])
            elif len(parts) in (8, 9):
                temp = float(parts[5])
            elif len(parts) == 1:
                continue
            else:
                continue
            if temp is not None:
                data.append({'year': year, 'temp': round(temp, 3)})

    # Bin to 5-year averages for chart
    binned = {}
    for d in data:
        bin_year = int(round(d['year'] / 5) * 5)
        if bin_year not in binned:
            binned[bin_year] = []
        binned[bin_year].append(d['temp'])

    result = []
    for year in sorted(binned.keys()):
        vals = binned[year]
        mean_temp = sum(vals) / len(vals)
        result.append({'year': year, 'temp': round(mean_temp, 3)})

    print(f"Büntgen tree rings: {len(data)} annual points → {len(result)} 5-year bins")
    return write_json('tree_rings_buentgen.json', result)


# =============================================================================
# 3. eVolv2k v4 — Volcanic eruptions (UPGRADE)
#    Source: evolv2k_v4.tab from PANGAEA
#    We want events from 500 BCE to 1000 CE with significant sulfur injection
# =============================================================================
def process_evolv2k():
    filepath = os.path.join(RAW, 'evolv2k_v4.tab')
    data = []
    with open(filepath) as f:
        lines = f.readlines()

    in_data = False
    for line in lines:
        if line.startswith('Eruption [a AD]'):
            in_data = True
            continue
        if not in_data:
            continue

        parts = line.strip().split('\t')
        if len(parts) < 11:
            continue

        try:
            year_ad = int(parts[0])
            vssi = float(parts[7]) if parts[7] else 0
            location = parts[10] if len(parts) > 10 else 'Unknown'
        except (ValueError, IndexError):
            continue

        if -500 <= year_ad <= 1000 and vssi >= 1.0:
            data.append({
                'year': year_ad,
                'sulfur': round(vssi, 2),
                'label': location if location != 'N/A' else 'Unknown'
            })

    data.sort(key=lambda d: d['year'])
    print(f"eVolv2k volcanic: {len(data)} significant events (VSSI >= 1 Tg, 500 BCE–1000 CE)")
    return write_json('volcanic_evolv2k.json', data)


# =============================================================================
# 4. Muigg/Tegel 2025 — Construction timber (NEW)
#    Source: timber_dataset.xlsx from Zenodo (doi:10.5281/zenodo.17090407)
#    20,397 dated wood samples with start_date, end_date
#    Bin by 25-year periods using felling date (end_date)
# =============================================================================
def process_timber():
    wb = openpyxl.load_workbook(os.path.join(RAW, 'timber_dataset.xlsx'), read_only=True)
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]

    end_date_idx = headers.index('end_date')
    felling_type_idx = headers.index('felling_date_type')

    dates = []
    for row in rows[1:]:
        end_date = row[end_date_idx]
        if end_date is not None:
            try:
                dates.append(int(end_date))
            except (ValueError, TypeError):
                continue

    # Bin by 25-year periods
    bin_width = 25
    bins = Counter()
    for d in dates:
        bin_start = (d // bin_width) * bin_width
        bins[bin_start] += 1

    result = []
    for bin_start in sorted(bins.keys()):
        mid = bin_start + bin_width / 2
        if -400 <= mid <= 800:
            result.append({'mid': mid, 'count': bins[bin_start]})

    print(f"Timber: {len(dates)} dated woods → {len(result)} 25-year bins")
    return write_json('timber_tegel.json', result)


# =============================================================================
# 5. Trentacoste — Cattle biometry (NEW)
#    Source: NItaly_Livestock_Metric_Data.csv from Zenodo (doi:10.5281/zenodo.6917159)
#    Compute log-size-index (LSI) using greatest length (GL) of metacarpals
#    Reference standard: modern cattle GL metacarpal ~195mm
# =============================================================================
def process_cattle():
    filepath = os.path.join(RAW, 'trentacoste_cattle.csv')

    # Collect GL measurements for cattle metacarpals by period
    period_data = defaultdict(list)
    period_map = {
        'Bronze Age': ('Bronze Age', -1200),
        'Bronze Age-Iron Age': ('Bronze Age/Iron Age', -900),
        'Iron Age': ('Iron Age', -500),
        'Roman': ('Roman', 100),
    }

    with open(filepath) as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['TaxonName'] != 'Cattle':
                continue
            period = row['Period']
            if period not in period_map:
                continue
            # Use GL (Greatest Length) for LSI calculation
            gl_str = row.get('GL', '')
            if gl_str and gl_str != 'NA':
                try:
                    gl = float(gl_str)
                    if 100 < gl < 300:  # sanity check for cattle metacarpal range
                        period_data[period].append(gl)
                except ValueError:
                    pass

    # If GL data is sparse, also use Bd (distal breadth) which is more common
    with open(filepath) as f:
        reader = csv.DictReader(f)
        bd_data = defaultdict(list)
        for row in reader:
            if row['TaxonName'] != 'Cattle':
                continue
            period = row['Period']
            if period not in period_map:
                continue
            bd_str = row.get('Bd', '')
            if bd_str and bd_str != 'NA':
                try:
                    bd = float(bd_str)
                    if 20 < bd < 100:
                        bd_data[period].append(bd)
                except ValueError:
                    pass

    # Use Bd (distal breadth) as it's more abundant — compute LSI relative to mean
    # Reference: use Iron Age mean as baseline (LSI = 0)
    all_bd = []
    for period in bd_data:
        all_bd.extend(bd_data[period])

    if 'Iron Age' in bd_data and bd_data['Iron Age']:
        reference = sum(bd_data['Iron Age']) / len(bd_data['Iron Age'])
    else:
        reference = sum(all_bd) / len(all_bd)

    import math
    result = []
    for period, (label, mid_year) in period_map.items():
        if period in bd_data and bd_data[period]:
            vals = bd_data[period]
            mean_bd = sum(vals) / len(vals)
            lsi = math.log10(mean_bd / reference)
            result.append({
                'period': label,
                'mid': mid_year,
                'lsi': round(lsi, 4),
                'mean_bd': round(mean_bd, 1),
                'n': len(vals)
            })

    result.sort(key=lambda d: d['mid'])
    for r in result:
        print(f"  {r['period']}: LSI={r['lsi']}, mean Bd={r['mean_bd']}mm, n={r['n']}")

    print(f"Cattle biometry: {len(result)} period means from {sum(len(v) for v in bd_data.values())} Bd measurements")
    return write_json('cattle_trentacoste.json', result)


# =============================================================================
# 6. Maddison 2023 — GDP per capita (NEW)
#    Source: mpd2023_web.xlsx from DataverseNL (doi:10.34894/INZBF2)
#    Extract ancient GDP per capita for Roman Empire core regions
# =============================================================================
def process_maddison():
    wb = openpyxl.load_workbook(
        os.path.join(RAW, 'maddison2023', 'mpd2023_web.xlsx'), read_only=True)
    ws = wb['Full data']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]

    # Columns: countrycode, country, region, year, gdppc, pop
    roman_codes = {
        'ITA': 'Italy', 'EGY': 'Egypt', 'GRC': 'Greece', 'TUR': 'Turkey',
        'ESP': 'Spain', 'FRA': 'France', 'GBR': 'Britain', 'IRQ': 'Iraq',
        'IRN': 'Iran', 'IND': 'India', 'CHN': 'China', 'JPN': 'Japan',
    }

    data = []
    for row in rows[1:]:
        code, country, region, year, gdppc, pop = row[:6]
        if code in roman_codes and year is not None and gdppc is not None:
            try:
                y = int(year)
                g = float(gdppc)
            except (ValueError, TypeError):
                continue
            if y <= 1500:
                data.append({
                    'year': y,
                    'gdp': round(g),
                    'country': roman_codes[code],
                    'code': code
                })

    data.sort(key=lambda d: (d['year'], d['country']))
    print(f"Maddison GDP: {len(data)} ancient data points across {len(set(d['country'] for d in data))} countries")
    return write_json('gdp_maddison.json', data)


# =============================================================================
# MAIN — process everything and write output
# =============================================================================
if __name__ == '__main__':
    print("=" * 60)
    print("Processing downloaded datasets → JSON files")
    print(f"Output directory: {SVELTE_DATA}")
    print("=" * 60)

    print("\n--- McConnell Lead ---")
    process_mcconnell()

    print("\n--- Büntgen Tree Rings ---")
    process_buentgen()

    print("\n--- eVolv2k Volcanic ---")
    process_evolv2k()

    print("\n--- Timber ---")
    process_timber()

    print("\n--- Cattle Biometry ---")
    process_cattle()

    print("\n--- Maddison GDP ---")
    process_maddison()

    print(f"\n{'=' * 60}")
    print("Done. JSON files written to svelte-app/src/lib/data/")
    print("Import them in datasets.js with: import data from './filename.json'")
    print(f"{'=' * 60}")
