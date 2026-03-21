#!/usr/bin/env python3
"""
Unified data processing pipeline: raw sources → Parquet cache → web-ready JSON.

All JSON outputs go to svelte-app/src/lib/data/ and are imported by datasets.js.
All parquet intermediates go to data/cache/ (gitignored).
See data/METHODOLOGY.md for the Met statistical methodology.

Usage:
    python data/process.py                  # process everything
    python data/process.py --skip-met       # skip Met (faster, no large CSV)
    python data/process.py --met-only       # only Met data
    python data/process.py --force          # rebuild Met parquet from raw CSV
    python data/process.py --analyze        # print Met date span analysis
"""
import sys
import json
import math
from pathlib import Path

import numpy as np
import polars as pl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).parent
RAW = ROOT / "raw"
CACHE = ROOT / "cache"
SVELTE = ROOT.parent / "svelte-app" / "src" / "lib" / "data"

CACHE.mkdir(exist_ok=True)


def write_json(filename: str, data) -> None:
    """Write compact JSON to the Svelte data directory."""
    path = SVELTE / filename
    with open(path, "w") as f:
        json.dump(data, f, separators=(",", ":"), ensure_ascii=False)
    size_kb = path.stat().st_size / 1024
    print(f"  → {filename} ({size_kb:.1f} KB)")


# ═══════════════════════════════════════════════════════════════════════════
# 1. McConnell 2018 — Lead emissions
#    Source: pnas.1721818115.sd01.xlsx, sheet "Fig. 3"
#    BP dates, 11-yr median filtered emissions (kt/a)
# ═══════════════════════════════════════════════════════════════════════════
def process_lead():
    import openpyxl

    wb = openpyxl.load_workbook(RAW / "pnas.1721818115.sd01.xlsx", read_only=True)
    rows = list(wb["Fig. 3"].iter_rows(values_only=True))

    records = [
        {"year": round(1950 - r[0], 1), "emissions": round(float(r[2]), 4)}
        for r in rows[7:]
        if r[0] is not None and r[2] is not None and r[2] != -0.999
    ]

    df = (
        pl.DataFrame(records)
        .filter(pl.col("year").is_between(-1200, 800))
        .sort("year")
    )
    df.write_parquet(CACHE / "lead-emissions.parquet")

    binned = (
        df
        .with_columns(((pl.col("year") / 5).round(0) * 5).cast(pl.Int32).alias("bin"))
        .group_by("bin")
        .agg(pl.col("emissions").mean().round(4))
        .sort("bin")
        .rename({"bin": "year"})
    )

    print(f"  Lead: {len(df)} annual → {len(binned)} 5-year bins")
    write_json("lead-emissions_mcconnell2018.json", binned.to_dicts())


# ═══════════════════════════════════════════════════════════════════════════
# 2. Büntgen 2011 — Tree ring temperature anomaly
#    Source: buentgen2011europe.txt from NOAA NCEI
#    Variable-width whitespace format with changing column layouts
# ═══════════════════════════════════════════════════════════════════════════
def process_tree_rings():
    lines = (RAW / "buentgen2011europe.txt").read_text().splitlines()

    records = []
    in_data = False
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("Year") and "TempJJA" in stripped:
            in_data = True
            continue
        if not in_data:
            continue
        parts = stripped.split()
        if not parts:
            break
        try:
            year = int(parts[0])
        except ValueError:
            break
        if year > 1000:
            break
        # Column layout (verified against raw file):
        #   4 cols: year | TempJJA | T- | T+
        #   7 cols: year | Precip | P- | P+ | TempJJA | T- | T+
        #   8 cols: year | Precip | P- | P+ | TempJJA | T- | T+ | Temp2006
        #   9 cols: year | Precip | P- | P+ | Precip2010 | TempJJA | T- | T+ | Temp2006
        n = len(parts)
        if n == 4:
            temp = float(parts[1])
        elif n in (7, 8):
            temp = float(parts[4])
        elif n == 9:
            temp = float(parts[5])
        elif n == 1:
            continue
        else:
            continue
        records.append({"year": year, "temp": round(temp, 3)})

    df = pl.DataFrame(records).sort("year")
    df.write_parquet(CACHE / "tree-rings.parquet")

    binned = (
        df
        .with_columns(((pl.col("year") / 5).round(0) * 5).cast(pl.Int32).alias("bin"))
        .group_by("bin")
        .agg(pl.col("temp").mean().round(3))
        .sort("bin")
        .rename({"bin": "year"})
    )

    print(f"  Tree rings: {len(df)} annual → {len(binned)} 5-year bins")
    write_json("tree-rings_buentgen2011.json", binned.to_dicts())


# ═══════════════════════════════════════════════════════════════════════════
# 3. eVolv2k v4 — Volcanic eruptions
#    Source: evolv2k_v4.tab from PANGAEA (tab-delimited with metadata header)
# ═══════════════════════════════════════════════════════════════════════════
def process_volcanic():
    lines = (RAW / "evolv2k_v4.tab").read_text().splitlines()

    records = []
    in_data = False
    for line in lines:
        if line.startswith("Eruption [a AD]"):
            in_data = True
            continue
        if not in_data:
            continue
        parts = line.strip().split("\t")
        if len(parts) < 11:
            continue
        try:
            year = int(parts[0])
            vssi = float(parts[7]) if parts[7] else 0
            location = parts[10] if len(parts) > 10 else "Unknown"
        except (ValueError, IndexError):
            continue
        if -500 <= year <= 1000 and vssi >= 1.0:
            records.append({
                "year": year,
                "sulfur": round(vssi, 2),
                "label": location if location != "N/A" else "Unknown",
            })

    df = pl.DataFrame(records).sort("year")
    df.write_parquet(CACHE / "volcanic-events.parquet")

    print(f"  Volcanic: {len(df)} events (VSSI ≥ 1 Tg, 500 BCE–1000 CE)")
    write_json("volcanic-events_evolv2k.json", df.to_dicts())


# ═══════════════════════════════════════════════════════════════════════════
# 4. Muigg/Tegel 2025 — Construction timber
#    Source: timber_dataset.xlsx from Zenodo
#    Filter to reliable felling dates, bin by 25-year periods
# ═══════════════════════════════════════════════════════════════════════════
def process_timber():
    df_raw = pl.read_excel(
        RAW / "timber_dataset.xlsx", sheet_name="Sheet1", engine="openpyxl"
    )

    df = (
        df_raw
        .filter(pl.col("felling_date_type").is_in(["exact felling date", "felling date interval"]))
        .filter(pl.col("end_date").is_not_null())
        .with_columns(pl.col("end_date").cast(pl.Int32))
    )
    df.write_parquet(CACHE / "timber-construction.parquet")

    skipped = len(df_raw) - len(df)
    binned = (
        df
        .with_columns((pl.col("end_date").floordiv(25) * 25).alias("bin_start"))
        .group_by("bin_start")
        .agg(pl.col("end_date").count().alias("count"))
        .with_columns((pl.col("bin_start") + 12.5).alias("mid"))
        .filter(pl.col("mid").is_between(-400, 800))
        .sort("mid")
        .select("mid", "count")
    )

    print(f"  Timber: {len(df)} reliable dates ({skipped} excluded) → {len(binned)} bins")
    write_json("timber-construction_tegel2025.json", binned.to_dicts())


# ═══════════════════════════════════════════════════════════════════════════
# 5. Trentacoste — Cattle biometry
#    Source: NItaly_Livestock_Metric_Data.csv from Zenodo
#    Relative LSI from distal breadth (Bd), Iron Age baseline
# ═══════════════════════════════════════════════════════════════════════════
PERIOD_MAP = {
    "Bronze Age": {"label": "Bronze Age", "mid": -1200},
    "Iron Age": {"label": "Iron Age", "mid": -500},
    "Roman": {"label": "Roman", "mid": 100},
}


def process_cattle():
    df = (
        pl.read_csv(RAW / "trentacoste_cattle.csv", null_values="NA", infer_schema_length=10000)
        .filter(pl.col("TaxonName") == "Cattle")
        .filter(pl.col("Period").is_in(list(PERIOD_MAP.keys())))
        .filter(pl.col("Bd").is_not_null())
        .with_columns(pl.col("Bd").cast(pl.Float64))
        .filter(pl.col("Bd").is_between(20, 100, closed="none"))
    )
    df.write_parquet(CACHE / "cattle-biometry.parquet")

    reference = df.filter(pl.col("Period") == "Iron Age")["Bd"].mean()

    result = (
        df
        .group_by("Period")
        .agg(
            pl.col("Bd").mean().round(1).alias("mean_bd"),
            pl.col("Bd").count().alias("n"),
        )
        .filter(pl.col("n") >= 3)
        .with_columns(
            (pl.col("mean_bd") / reference).log(base=10).round(4).alias("lsi"),
            pl.col("Period").replace_strict(
                {k: v["label"] for k, v in PERIOD_MAP.items()}
            ).alias("period"),
            pl.col("Period").replace_strict(
                {k: v["mid"] for k, v in PERIOD_MAP.items()}
            ).cast(pl.Int32).alias("mid"),
        )
        .sort("mid")
        .select("period", "mid", "lsi", "mean_bd", "n")
    )

    for r in result.iter_rows(named=True):
        print(f"    {r['period']}: LSI={r['lsi']}, Bd={r['mean_bd']}mm, n={r['n']}")
    print(f"  Cattle: {len(result)} periods from {len(df)} measurements")
    write_json("cattle-biometry_trentacoste.json", result.to_dicts())


# ═══════════════════════════════════════════════════════════════════════════
# 6. Maddison 2023 — GDP per capita
#    Source: mpd2023_web.xlsx from DataverseNL
#    Pre-1500 GDP for Mediterranean + comparison economies
# ═══════════════════════════════════════════════════════════════════════════
COUNTRY_CODES = {
    "ITA": "Italy", "EGY": "Egypt", "GRC": "Greece", "TUR": "Turkey",
    "ESP": "Spain", "FRA": "France", "GBR": "Britain", "IRQ": "Iraq",
    "IRN": "Iran", "IND": "India", "CHN": "China", "JPN": "Japan",
}


def process_gdp():
    df_raw = pl.read_excel(
        RAW / "maddison2023" / "mpd2023_web.xlsx",
        sheet_name="Full data", engine="openpyxl",
    )

    df = (
        df_raw
        .filter(pl.col("countrycode").is_in(list(COUNTRY_CODES.keys())))
        .filter(pl.col("year").is_not_null(), pl.col("gdppc").is_not_null())
        .with_columns(
            pl.col("year").cast(pl.Int32),
            pl.col("gdppc").cast(pl.Float64).round(0).cast(pl.Int32).alias("gdp"),
            pl.col("countrycode").replace_strict(COUNTRY_CODES).alias("country"),
        )
        .filter(pl.col("year") <= 1500)
        .sort("year", "country")
        .select("year", "gdp", "country", pl.col("countrycode").alias("code"))
    )
    df.write_parquet(CACHE / "gdp-per-capita.parquet")

    print(f"  GDP: {len(df)} data points across {df['country'].n_unique()} countries")
    write_json("gdp-per-capita_maddison2023.json", df.to_dicts())


# ═══════════════════════════════════════════════════════════════════════════
# 7. Shipwrecks — Strauss/OxREP
#    Source: pre-processed CSVs in data/processed/ (from src/ingest_shipwrecks.py)
#    Cache full-resolution data, export binned versions as JSON
# ═══════════════════════════════════════════════════════════════════════════
def process_shipwrecks():
    processed = ROOT / "processed"
    clean_csv = processed / "shipwrecks_clean.csv"

    if clean_csv.exists():
        df = pl.read_csv(clean_csv, infer_schema_length=10000)
        df.write_parquet(CACHE / "shipwrecks.parquet")
        print(f"  Shipwrecks: cached {len(df)} records")

    for suffix, name in [("25yr", "shipwrecks-25yr_strauss.json"),
                         ("50yr", "shipwrecks-50yr_strauss.json")]:
        csv_path = processed / f"shipwrecks_binned_{suffix}.csv"
        if csv_path.exists():
            df = pl.read_csv(csv_path)
            write_json(name, df.to_dicts())


# ═══════════════════════════════════════════════════════════════════════════
# 8. Met Open Access — Museum artifact data
#    Source: raw/met-artifact-data/MetObjects.csv (303 MB)
#    Pipeline: CSV → Parquet cache → dual-kernel probabilistic binning → JSON
#    See METHODOLOGY.md for the full statistical model.
# ═══════════════════════════════════════════════════════════════════════════
MET_BIN_WIDTH = 25
MET_MAX_YEAR = 1500
Z_95 = 1.96

MET_COMPOSITE_CLASSES = [
    "Vases", "Ceramics", "Glass", "Sculpture", "Bronzes",
    "Coins", "Jewelry", "Gold and Silver", "Terracottas", "Metalwork",
]

MET_FOCUS_DEPTS = [
    "Greek and Roman Art", "Egyptian Art", "Islamic Art",
    "Medieval Art", "Ancient Near Eastern Art", "Asian Art",
    "European Sculpture and Decorative Arts", "Arms and Armor",
]

MET_FOCUS_CLS = [
    "Vases", "Ceramics", "Glass", "Sculpture", "Bronzes",
    "Coins", "Jewelry", "Gold and Silver", "Gems",
    "Ceramics-Pottery", "Terracottas", "Stone Sculpture",
    "Textiles-Woven", "Metalwork", "Metalwork-Silver", "Paintings",
]

MET_FOCUS_CULTURES = [
    "Greek, Attic", "Roman", "Cypriot", "Etruscan", "Greek",
    "Coptic", "Sasanian", "Byzantine", "Frankish",
    "China", "Japan", "Iran",
]


def met_load_parquet(force=False):
    """Load Met CSV → Parquet cache. Returns a pandas DataFrame for numpy interop."""
    import pandas as pd

    parquet_path = CACHE / "met-objects.parquet"
    raw_csv = RAW / "met-artifact-data" / "MetObjects.csv"

    if parquet_path.exists() and not force:
        print(f"  Loading cached parquet ({parquet_path.stat().st_size / 1e6:.0f} MB)")
        df = pd.read_parquet(parquet_path)
        print(f"  {len(df):,} rows, {len(df.columns)} columns")
        return df

    print(f"  Reading raw CSV: {raw_csv}")
    df = pd.read_csv(raw_csv, encoding="utf-8-sig", low_memory=False)
    print(f"  Raw: {len(df):,} rows, {len(df.columns)} columns")

    for col in ["Is Highlight", "Is Timeline Work", "Is Public Domain"]:
        df[col] = df[col].map({"True": True, "False": False, True: True, False: False}).astype("boolean")
    for col in ["Object ID", "Object Begin Date", "Object End Date"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
    df["AccessionYear"] = pd.to_numeric(df["AccessionYear"], errors="coerce").astype("Int64")
    str_cols = [c for c in df.columns if df[c].dtype == object]
    for col in str_cols:
        df[col] = df[col].astype("string")

    df.to_parquet(parquet_path, engine="pyarrow", index=False)
    print(f"  → Wrote parquet ({parquet_path.stat().st_size / 1e6:.1f} MB)")
    return df


def _norm_cdf(x):
    """Vectorized standard normal CDF (Abramowitz & Stegun, ~1e-7 accuracy)."""
    a = np.abs(x)
    t = 1.0 / (1.0 + 0.2316419 * a)
    d = 0.3989422804014327
    poly = t * (0.319381530 + t * (-0.356563782 + t * (
        1.781477937 + t * (-1.821255978 + t * 1.330274429))))
    p = d * np.exp(-0.5 * x * x) * poly
    return np.where(x >= 0, 1.0 - p, p)


def _compute_bin_stats(begins, ends, bin_starts, bin_width=MET_BIN_WIDTH):
    """Vectorized dual-kernel per-bin statistics (uniform + Gaussian)."""
    n_bins = len(bin_starts)
    if len(begins) == 0:
        z = np.zeros(n_bins)
        return {"u_mean": z, "u_var": z.copy(), "g_mean": z.copy(),
                "g_var": z.copy(), "n_objects": np.zeros(n_bins, dtype=int)}

    eff_ends = np.where(ends <= begins, begins + 1.0, ends)
    spans = eff_ends - begins
    centers = (begins + eff_ends) / 2.0
    sigmas = np.maximum(spans / 4.0, 0.5)

    z_lo = (begins - centers) / sigmas
    z_hi = (eff_ends - centers) / sigmas
    Z = np.maximum(_norm_cdf(z_hi) - _norm_cdf(z_lo), 1e-12)
    tiny = spans < 4

    u_mean = np.zeros(n_bins)
    u_var = np.zeros(n_bins)
    g_mean = np.zeros(n_bins)
    g_var = np.zeros(n_bins)
    n_objects = np.zeros(n_bins, dtype=int)

    for i in range(n_bins):
        bs, be = bin_starts[i], bin_starts[i] + bin_width
        ol_lo = np.maximum(begins, bs)
        ol_hi = np.minimum(eff_ends, be)
        overlap = np.maximum(0.0, ol_hi - ol_lo)
        p_u = overlap / spans

        z1 = (bs - centers) / sigmas
        z2 = (be - centers) / sigmas
        p_g = np.clip((_norm_cdf(z2) - _norm_cdf(z1)) / Z, 0.0, 1.0)
        p_g = np.where(tiny, p_u, p_g)
        p_g = np.where(overlap > 0, p_g, 0.0)

        has_overlap = overlap > 0
        u_mean[i] = np.sum(p_u)
        u_var[i] = np.sum(p_u * (1.0 - p_u))
        g_mean[i] = np.sum(p_g)
        g_var[i] = np.sum(p_g * (1.0 - p_g))
        n_objects[i] = np.sum(has_overlap)

    return {"u_mean": u_mean, "u_var": u_var, "g_mean": g_mean,
            "g_var": g_var, "n_objects": n_objects}


def _stats_to_series(stats, bin_mids):
    """Convert bin stats to JSON-ready records with dual-kernel CIs."""
    series = []
    for i, mid in enumerate(bin_mids):
        u_m, g_m = stats["u_mean"][i], stats["g_mean"][i]
        n = int(stats["n_objects"][i])
        if u_m < 0.01 and n == 0:
            continue
        u_ci = Z_95 * np.sqrt(stats["u_var"][i])
        g_ci = Z_95 * np.sqrt(stats["g_var"][i])
        series.append({
            "mid": mid, "count": round(u_m, 2),
            "lo": round(max(0, u_m - u_ci), 2), "hi": round(u_m + u_ci, 2),
            "count_g": round(g_m, 2),
            "lo_g": round(max(0, g_m - g_ci), 2), "hi_g": round(g_m + g_ci, 2),
            "n_objects": n,
        })
    return series


def _met_build_timelines(df, year_range, bin_width=MET_BIN_WIDTH, groupby_col=None):
    """Build dual-kernel 25-year bin series from pandas DataFrame."""
    mask = (
        df["Object Begin Date"].notna() & df["Object End Date"].notna()
        & (df["Object Begin Date"] <= year_range[1])
        & (df["Object End Date"] >= year_range[0])
        & (df["Object Begin Date"] >= -3000) & (df["Object End Date"] <= 3000)
    )
    filt = df[mask]
    begins = filt["Object Begin Date"].to_numpy(dtype="float64")
    ends = filt["Object End Date"].to_numpy(dtype="float64")

    bin_starts = np.arange(
        (year_range[0] // bin_width) * bin_width,
        year_range[1] + bin_width, bin_width, dtype=float)
    bin_mids = bin_starts + bin_width / 2.0

    if groupby_col is None:
        return _stats_to_series(_compute_bin_stats(begins, ends, bin_starts, bin_width), bin_mids)

    groups_raw = filt[groupby_col].fillna("").astype(str).str.strip().values
    unique = sorted(set(g for g in groups_raw if g))
    result = {}
    for g in unique:
        m = groups_raw == g
        s = _stats_to_series(_compute_bin_stats(begins[m], ends[m], bin_starts, bin_width), bin_mids)
        if s:
            result[g] = s
    return result


def _met_composite(per_class_series, classes=MET_COMPOSITE_CLASSES):
    """Equal-weight composite index with propagated CIs."""
    valid = {}
    for cls in classes:
        if cls not in per_class_series:
            continue
        s = per_class_series[cls]
        u_pk = max((r["count"] for r in s), default=0)
        g_pk = max((r["count_g"] for r in s), default=0)
        if u_pk < 5:
            continue
        valid[cls] = {"lookup": {r["mid"]: r for r in s}, "u_pk": u_pk, "g_pk": max(g_pk, 1)}

    if not valid:
        return {"uniform": [], "gaussian": [], "classes_used": []}

    all_mids = sorted({m for v in valid.values() for m in v["lookup"]})
    n = len(valid)
    u_out, g_out = [], []

    for mid in all_mids:
        u_norms, u_vars, g_norms, g_vars = [], [], [], []
        for vc in valid.values():
            r = vc["lookup"].get(mid)
            if r:
                u_norms.append(r["count"] / vc["u_pk"])
                u_vars.append(((r["hi"] - r["count"]) / Z_95) ** 2 / vc["u_pk"] ** 2)
                g_norms.append(r["count_g"] / vc["g_pk"])
                g_vars.append(((r["hi_g"] - r["count_g"]) / Z_95) ** 2 / vc["g_pk"] ** 2)
            else:
                u_norms.append(0); u_vars.append(0)
                g_norms.append(0); g_vars.append(0)

        u_m = sum(u_norms) / n
        g_m = sum(g_norms) / n
        u_ci = Z_95 * np.sqrt(sum(u_vars) / n ** 2)
        g_ci = Z_95 * np.sqrt(sum(g_vars) / n ** 2)
        u_out.append({"mid": mid, "index": round(u_m, 4),
                       "lo": round(max(0, u_m - u_ci), 4), "hi": round(u_m + u_ci, 4)})
        g_out.append({"mid": mid, "index": round(g_m, 4),
                       "lo": round(max(0, g_m - g_ci), 4), "hi": round(g_m + g_ci, 4)})

    return {"uniform": u_out, "gaussian": g_out, "classes_used": list(valid.keys())}


def process_met(force=False, analyze=False):
    """Full Met pipeline: parquet cache → dual-kernel binning → JSON exports."""
    import pandas as pd

    df = met_load_parquet(force=force)

    if analyze:
        _met_analyze(df)

    yr = (-1000, MET_MAX_YEAR)

    print("  Building timelines...")
    write_json("met-timeline-25yr_metmuseum.json", _met_build_timelines(df, yr))

    dept = _met_build_timelines(df, yr, groupby_col="Department")
    write_json("met-by-department_metmuseum.json",
               {d: dept[d] for d in MET_FOCUS_DEPTS if d in dept})

    cls = _met_build_timelines(df, yr, groupby_col="Classification")
    cls_json = {c: cls[c] for c in MET_FOCUS_CLS if c in cls}
    write_json("met-by-classification_metmuseum.json", cls_json)

    cul = _met_build_timelines(df, yr, groupby_col="Culture")
    write_json("met-by-culture_metmuseum.json",
               {c: cul[c] for c in MET_FOCUS_CULTURES if c in cul})

    gr_mask = df["Department"] == "Greek and Roman Art"
    gr = _met_build_timelines(df[gr_mask], (-1000, 600), groupby_col="Classification")
    write_json("met-greek-roman_metmuseum.json",
               {c: s for c, s in gr.items() if sum(r["count"] for r in s) >= 20})

    write_json("met-composite-index_metmuseum.json", _met_composite(cls_json))


def _met_analyze(df):
    """Print date span distributions (optional diagnostic)."""
    import pandas as pd

    anc = df[
        df["Object Begin Date"].notna() & df["Object End Date"].notna()
        & (df["Object Begin Date"] >= -3000) & (df["Object End Date"] <= 2000)
    ].copy()
    anc["span"] = anc["Object End Date"] - anc["Object Begin Date"]

    print(f"\n  Date span analysis — {len(anc):,} objects")
    for p in [25, 50, 75, 90, 95]:
        print(f"    p{p}: {anc['span'].quantile(p/100):.0f} yr")
    print(f"    mean: {anc['span'].mean():.0f} yr")


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    skip_met = "--skip-met" in sys.argv
    met_only = "--met-only" in sys.argv
    force = "--force" in sys.argv
    analyze = "--analyze" in sys.argv

    print("=" * 60)
    print("Data pipeline: raw → parquet cache → web JSON")
    print("=" * 60)

    if not met_only:
        print("\n--- Lead emissions (McConnell 2018) ---")
        process_lead()
        print("\n--- Tree rings (Büntgen 2011) ---")
        process_tree_rings()
        print("\n--- Volcanic events (eVolv2k) ---")
        process_volcanic()
        print("\n--- Timber construction (Tegel 2025) ---")
        process_timber()
        print("\n--- Cattle biometry (Trentacoste) ---")
        process_cattle()
        print("\n--- GDP per capita (Maddison 2023) ---")
        process_gdp()
        print("\n--- Shipwrecks (Strauss/OxREP) ---")
        process_shipwrecks()

    if not skip_met:
        print("\n--- Met Open Access ---")
        process_met(force=force, analyze=analyze)

    print(f"\n{'=' * 60}")
    print("Done. JSON written to svelte-app/src/lib/data/")
    print(f"{'=' * 60}")
